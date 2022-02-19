from openpyxl import Workbook

wb = Workbook()
#ws = wb.active

ws1=wb.create_sheet("avinash details")
ws2=wb.create_sheet("bharath details")
ws3=wb.create_sheet("visu details")

ws1["A1"]="name"
ws1["B1"]="location"
ws1["C1"]="profession"
ws1["A2"]="avinash"
ws1["B2"]="nellore"
ws1["C2"]="engineer"

ws2["A1"]="name"
ws2["B1"]="location"
ws2["C1"]="profession"
ws2["A2"]="bharath"
ws2["B2"]="ongole"
ws2["C2"]="engineer"

ws3["A1"]="name"
ws3["B1"]="location"
ws3["C1"]="profession"
ws3["A2"]="visu"
ws3["B2"]="kovur"
ws3["C2"]="engineer"

wb.save("team details.xlsx")