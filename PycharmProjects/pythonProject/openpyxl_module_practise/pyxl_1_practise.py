from openpyxl import Workbook

wb = Workbook()
#ws = wb.active

ws1 = wb.create_sheet("Avinash DSR")
ws2 = wb.create_sheet("Bharath DSR")
ws3 = wb.create_sheet("Visu DSR")

#ws1["A1", "B1", "C1", "D1"]= ["DATE", "TOPICS", "MORNING", "EVENING"]
ws1["A1"] = "Date"
ws1["B1"] = "Topics"
ws1["C1"] = "Morning"
ws1["D1"] = "Evening"
ws2["A1"] = "Date"
ws1["A2"] = "14-12-2021"

wb.save("15-12-2021 DSR.xlsx")