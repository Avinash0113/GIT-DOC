# import openpyxl
#
# wb = openpyxl.Workbook()
# #ws = wb.active
#
# ws1 = wb.create_sheet("Avinash_dsr")
# ws2 = wb.create_sheet("Bharath_dsr")
# ws3 = wb.create_sheet("Visu_dsr")
#
# # ws.title = "Weekly_DSR"
# # del ws
# # print(wb.sheetnames)
#
# ws1["A1"] = "Date"
# ws1["B1"] = "Topics"
# ws1["C1"] = "Programs"
# ws1["A2"] = "OOPs"
# # ws2["A1","B1","C1"] = "Date","Topics","Programs"
# # ws3["A1","B1","C1"] = "Date","Topics","Programs"
#
# wb.save("DSR.xlsx")


import openpyxl

wb_obj = openpyxl.Workbook()

sheet_obj = wb_obj.active

daily_status_data = (
    ("27-08-2021", "practicing_examples", "prepare for interview"),
    ("27-08-2021", "practicing_examples", "prepare for interview")
    )

for data in daily_status_data:
    sheet_obj.append(data)

wb_obj.save("DSR.xlsx")
