from openpyxl import Workbook

wb = Workbook()
ws = wb.active

ws1 = wb.create_sheet("Avinash DSR")
ws2 = wb.create_sheet("Bharath DSR")
ws3 = wb.create_sheet("Visu DSR")

ws1["A1"] = "Date"
ws1["B1"] = "Today topics"
ws1["A2"] = "14-12-2021"
ws1["A3"] = "16-12-2021"
ws1["B2"] = "oops"
ws1["B3"] = "oops"
ws2["A1"] = "date"
ws2["b1"] = "topics"

wb.save("DSR.xlsx")