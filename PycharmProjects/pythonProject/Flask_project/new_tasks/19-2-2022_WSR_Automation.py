import openpyxl

wb=openpyxl.Workbook()
sheet_1=wb.create_sheet("mahesh")
sheet_2=wb.create_sheet("visu")
sheet_3=wb.create_sheet("bharath")
sheet_4=wb.create_sheet("avinash")

rows= (("date","topics","practice"),
       ("19-2-2022","oops","programs"))
for row in rows:
    sheet_1.append(row)
wb.save("DSR.xlsx")





